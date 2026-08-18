# -*- coding: utf-8 -*-
"""将《应急装备数量统计（内部）》基准表导入 equipment_benchmark 汇总表。

基准表 Sheet1 同时包含左侧、右侧两块统计面板，左侧是卫星电话、370MHz、北斗，
右侧是便携站、指挥车、叫应终端、布控球、无人机。本模块按“设备类型+统计单位+口径”
读取各行数字，并把市局配发/区局自购、配置总数/可用数量等口径一并保存。
"""

import argparse
import re
from pathlib import Path

import openpyxl

from equipment.db import get_connection

# 基准表固定结构
BENCHMARK_SHEET = "Sheet1"
DATA_START_ROW = 10
DATA_END_ROW = 26
SKIP_KEYS = {"", "市/区", "总数："}

UPSERT_SQL = """
INSERT INTO equipment_benchmark (
    device_type, district, metric, count,
    source_file, source_sheet, source_row, updated_at
) VALUES (
    :device_type, :district, :metric, :count,
    :source_file, :source_sheet, :source_row, datetime('now', 'localtime')
)
ON CONFLICT(device_type, district, metric) DO UPDATE SET
    count = excluded.count,
    source_file = excluded.source_file,
    source_sheet = excluded.source_sheet,
    source_row = excluded.source_row,
    updated_at = datetime('now', 'localtime')
"""


def _to_int(value):
    """取单元格开头的数字；处理“1(河西消防使用)”这类附带说明。"""
    if value is None:
        return 0
    match = re.match(r"\s*(\d+)", str(value))
    return int(match.group(1)) if match else 0


def _row_items(row, source_file):
    """将基准表一行的各口径转为待入库记录，数量为 0 的口径不写入。"""
    handset_grant = _to_int(row[1])
    handset_purchase = _to_int(row[2])
    handset_total = handset_grant + handset_purchase
    m370_grant = _to_int(row[4])
    m370_purchase = _to_int(row[5])
    m370_config = m370_grant + m370_purchase
    m370_available = _to_int(row[6])

    candidates = [
        ("卫星电话（手持）", "市局配发", handset_grant),
        ("卫星电话（手持）", "区局自购", handset_purchase),
        ("卫星电话（手持）", "总数", handset_total),
        ("卫星电话（固移）", "总数", _to_int(row[3])),
        ("370MHz手持终端", "市局配发", m370_grant),
        ("370MHz手持终端", "区局自购", m370_purchase),
        ("370MHz手持终端", "配置总数", m370_config),
        ("370MHz手持终端", "可用数量", m370_available),
        ("北斗终端", "总数", _to_int(row[7])),
        ("卫星便携站", "总数", _to_int(row[9])),
        ("通信指挥车", "总数", _to_int(row[10])),
        ("叫应终端", "总数", _to_int(row[11])),
        ("布控球", "总数", _to_int(row[12])),
        ("无人机", "总数", _to_int(row[13])),
    ]
    return [
        (device_type, metric, count)
        for device_type, metric, count in candidates
        if count > 0
    ]


def parse_benchmark_rows(xlsx_path):
    """读取基准表统计行，返回原始记录列表。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if BENCHMARK_SHEET not in wb.sheetnames:
        raise ValueError(f"工作簿中缺少「{BENCHMARK_SHEET}」工作表")
    ws = wb[BENCHMARK_SHEET]
    source_file = Path(xlsx_path).name

    raw_rows = []
    for row_idx in range(DATA_START_ROW, DATA_END_ROW + 1):
        row = [
            ws.cell(row=row_idx, column=col).value
            for col in range(1, 15)
        ]
        if not any(value is not None and str(value).strip() for value in row):
            continue
        district = str(row[0]).strip()
        if district in SKIP_KEYS:
            continue
        if len(row) <= 13:
            raise ValueError(f"第 {row_idx} 行缺少无人机列，基准表结构不完整")
        for device_type, metric, count in _row_items(row, source_file):
            raw_rows.append(
                {
                    "device_type": device_type,
                    "district": district,
                    "metric": metric,
                    "count": count,
                    "source_file": source_file,
                    "source_sheet": BENCHMARK_SHEET,
                    "source_row": row_idx,
                }
            )
    if not raw_rows:
        raise ValueError("基准表统计区未读取到有效数据")
    return raw_rows


def ingest_benchmark_xlsx(xlsx_path, conn):
    """导入基准表汇总到 equipment_benchmark，返回统计。"""
    rows = parse_benchmark_rows(xlsx_path)
    stats = {"total": 0, "types": {}, "districts": []}
    seen_districts = set()

    for item in rows:
        conn.execute(UPSERT_SQL, item)
        stats["total"] += 1
        stats["types"][item["device_type"]] = (
            stats["types"].get(item["device_type"], 0) + 1
        )
        seen_districts.add(item["district"])

    conn.commit()
    stats["districts"] = sorted(seen_districts)
    return stats


def rebuild_benchmark(conn):
    """清空基准汇总表，用于基准表更新后重建。"""
    conn.execute("DELETE FROM equipment_benchmark")
    conn.commit()


def main():
    parser = argparse.ArgumentParser(
        description="将《应急装备数量统计（内部）》xlsx 导入 equipment_benchmark 汇总表"
    )
    parser.add_argument("xlsx", help="基准统计表 Excel 文件路径")
    parser.add_argument("--db", default=None, help="数据库路径（默认 data/equipment.db）")
    parser.add_argument(
        "--rebuild",
        action="store_true",
        help="先清空基准汇总表再导入（用于基准表更新后重建）",
    )
    args = parser.parse_args()

    conn = get_connection(args.db)
    if args.rebuild:
        rebuild_benchmark(conn)
    stats = ingest_benchmark_xlsx(args.xlsx, conn)
    print(f"导入完成：基准汇总 {stats['total']} 条，覆盖 {len(stats['districts'])} 个统计单位")
    for device_type, count in sorted(stats["types"].items()):
        print(f"  - {device_type}: {count} 条")
    conn.close()


if __name__ == "__main__":
    main()
