# -*- coding: utf-8 -*-
"""将《通信保障重点村应急通信方式统计》xlsx 导入风险村表。

风险村台账是全量台账的一个特殊范围。本脚本读取台账明细（区/街镇/村名），
逐村解析为与 devices 表对齐的标准键（district/town/unit_name），
写入 risk_villages 标准表，并将台账原始写法写入 risk_village_aliases 别名表。
联系人、电话、灾害类型等字段不入表。
"""

import argparse
from pathlib import Path

import openpyxl

from equipment.db import get_connection
from equipment.normalize import (
    RISK_VILLAGE_ALIASES,
    clean_risk_alias,
    clean_risk_district,
    clean_risk_town,
    clean_risk_village_name,
    cn_to_arabic,
    normalize_unit_name,
)

# 风险村台账工作表名与必要列
RISK_SHEET = "风险村台账"
REQUIRED_COLUMNS = {"区", "街镇", "通信保障重点村名称"}

UPSERT_VILLAGE_SQL = """
INSERT INTO risk_villages (district, town, village_name, source_row, created_at)
VALUES (:district, :town, :village_name, :source_row, datetime('now', 'localtime'))
ON CONFLICT(district, town, village_name) DO UPDATE SET
    source_row = excluded.source_row
"""

UPSERT_ALIAS_SQL = """
INSERT INTO risk_village_aliases (risk_village_id, alias_town, alias_name)
VALUES (:risk_village_id, :alias_town, :alias_name)
ON CONFLICT(alias_town, alias_name) DO UPDATE SET
    risk_village_id = excluded.risk_village_id
"""


def _cell_text(value):
    """单元格转文本并去除换行等空白。"""
    if value is None:
        return ""
    return str(value).replace("\n", "").replace(" ", "").replace("\u3000", "")


def parse_risk_sheet(xlsx_path):
    """读取风险村台账，返回原始行列表 [(区, 街镇, 村名)]。

    区/街镇为合并单元格，空单元格向下沿用前一行值。
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if RISK_SHEET not in wb.sheetnames:
        raise ValueError(f"工作簿中缺少「{RISK_SHEET}」工作表")
    ws = wb[RISK_SHEET]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"「{RISK_SHEET}」工作表为空")

    header = [_cell_text(c) for c in rows[0]]
    col_idx = {name: header.index(name) for name in REQUIRED_COLUMNS if name in header}
    missing = REQUIRED_COLUMNS - set(col_idx)
    if missing:
        raise ValueError(f"「{RISK_SHEET}」缺少列：{sorted(missing)}")

    raw_rows = []
    cur_district = ""
    cur_town = ""
    for row_idx, row in enumerate(rows[1:], start=2):
        if not any(c is not None and str(c).strip() != "" for c in row):
            continue
        district = _cell_text(row[col_idx["区"]])
        town = _cell_text(row[col_idx["街镇"]])
        name = _cell_text(row[col_idx["通信保障重点村名称"]])
        if district:
            cur_district = district
        if town:
            cur_town = town
        if not cur_district or not cur_town or not name:
            raise ValueError(f"第 {row_idx} 行区/街镇/村名不完整：{district}/{town}/{name}")
        raw_rows.append((cur_district, cur_town, name, row_idx))
    return raw_rows


def _candidate_units(conn, district, town):
    """查询同区同街镇的设备使用单位（去重，排除空单位）。"""
    return [
        row["unit_name"]
        for row in conn.execute(
            """
            SELECT DISTINCT unit_name FROM devices
            WHERE district = ? AND town = ?
              AND unit_name IS NOT NULL AND unit_name != ''
            """,
            (district, town),
        )
    ]


def _core_name(name):
    """去掉村/村委会尾缀，得到核心名（如 芳峪村 -> 芳峪、十三街村委会 -> 十三街）。"""
    for suffix in ("村委会", "村"):
        if name.endswith(suffix) and len(name) > len(suffix):
            return name[: -len(suffix)]
    return name


def resolve_risk_village(conn, raw_district, raw_town, raw_name):
    """将台账原文解析为 devices 中的标准键。

    返回 (district, town, village_name, alias_town, alias_name)；
    无法唯一匹配时返回 None。
    """
    district = clean_risk_district(raw_district)
    town = clean_risk_town(raw_town)
    clean_name = clean_risk_village_name(raw_name)
    alias_town = clean_risk_alias(raw_town)
    alias_name = clean_risk_alias(raw_name)

    candidates = _candidate_units(conn, district, town)
    # 1) 手工别名表：加字/别字村名直接映射到标准单位名，须存在于候选单位中
    std_alias = RISK_VILLAGE_ALIASES.get(raw_name)
    if std_alias:
        if std_alias in candidates:
            return district, town, std_alias, alias_town, alias_name
        return None

    # 2) 核心名精确相等（如 芳峪村 -> 芳峪村委会、十三街村 -> 十三街村委会）
    #    两侧都转阿拉伯数字，解决“柒街/七街”“拾叁街/十三街”写法差异
    core = _core_name(clean_name)
    normed_candidates = []
    for unit_name in candidates:
        norm = cn_to_arabic(normalize_unit_name(unit_name) or "")
        normed_candidates.append((unit_name, _core_name(norm)))
    for unit_name, norm_core in normed_candidates:
        if norm_core == core:
            return district, town, unit_name, alias_town, alias_name
    # 3) 核心名前缀匹配（如 凌沿村 -> 凌沿庄村委会、于台村 -> 于台子村委会）
    #    仅允许“长名以短名为前缀”，避免 芳峪村 误配 小穿芳峪村委会
    for unit_name, norm_core in normed_candidates:
        if len(core) >= 2 and (norm_core.startswith(core) or core.startswith(norm_core)):
            return district, town, unit_name, alias_town, alias_name

    # 4) 街镇级单位特例：天津铁厂街道（unit_name 为空，村名即街镇名）
    if town == clean_name:
        hit = conn.execute(
            """
            SELECT 1 FROM devices
            WHERE district = ? AND town = ?
              AND (unit_name IS NULL OR unit_name = '')
            LIMIT 1
            """,
            (district, town),
        ).fetchone()
        if hit:
            return district, town, town, alias_town, alias_name

    return None


def ingest_risk_xlsx(xlsx_path, conn):
    """导入风险村台账，返回统计。"""
    raw_rows = parse_risk_sheet(xlsx_path)
    stats = {"total": 0, "failed": [], "matched_units": 0}
    seen = set()

    for raw_district, raw_town, raw_name, source_row in raw_rows:
        resolved = resolve_risk_village(conn, raw_district, raw_town, raw_name)
        if resolved is None:
            stats["failed"].append((source_row, raw_district, raw_town, raw_name))
            continue

        district, town, village_name, alias_town, alias_name = resolved
        key = (district, town, village_name)
        if key in seen:
            # 台账同一标准村重复出现时只保留一行，避免标准表唯一键冲突
            continue
        seen.add(key)

        existing = conn.execute(
            """
            SELECT id FROM risk_villages
            WHERE district = ? AND town = ? AND village_name = ?
            """,
            (district, town, village_name),
        ).fetchone()
        if existing:
            village_id = existing["id"]
            conn.execute(
                "UPDATE risk_villages SET source_row = ? WHERE id = ?",
                (source_row, village_id),
            )
        else:
            village_id = conn.execute(
                UPSERT_VILLAGE_SQL,
                {
                    "district": district,
                    "town": town,
                    "village_name": village_name,
                    "source_row": source_row,
                },
            ).lastrowid
        # 先清理该村旧别名，再写入最新原文写法
        conn.execute(
            "DELETE FROM risk_village_aliases WHERE risk_village_id = ?",
            (village_id,),
        )
        conn.execute(
            UPSERT_ALIAS_SQL,
            {
                "risk_village_id": village_id,
                "alias_town": alias_town,
                "alias_name": alias_name,
            },
        )
        stats["total"] += 1

    conn.commit()
    stats["matched_units"] = conn.execute(
        """
        SELECT COUNT(*) FROM risk_villages rv
        JOIN devices d
          ON d.district = rv.district AND d.town = rv.town
         AND (COALESCE(d.unit_name, '') = rv.village_name
              OR (COALESCE(d.unit_name, '') = '' AND d.town = rv.village_name))
        """
    ).fetchone()[0]
    return stats


def rebuild_risk(conn):
    """清空风险村两张表（先删别名再删标准表，避免外键悬挂）。"""
    conn.execute("DELETE FROM risk_village_aliases")
    conn.execute("DELETE FROM risk_villages")
    conn.commit()


def main():
    parser = argparse.ArgumentParser(
        description="将风险村台账 xlsx 导入 risk_villages / risk_village_aliases 表"
    )
    parser.add_argument("xlsx", help="风险村台账 Excel 文件路径")
    parser.add_argument("--db", default=None, help="数据库路径（默认 data/equipment.db）")
    parser.add_argument(
        "--rebuild",
        action="store_true",
        help="先清空风险村表再导入（用于台账更新后重建）",
    )
    args = parser.parse_args()

    conn = get_connection(args.db)
    if args.rebuild:
        rebuild_risk(conn)
    stats = ingest_risk_xlsx(args.xlsx, conn)

    print(f"导入完成：风险村 {stats['total']} 个，命中全量台账设备单位 {stats['matched_units']} 个")
    if stats["failed"]:
        print(f"以下 {len(stats['failed'])} 行未能匹配到全量台账，请人工核对：")
        for row in stats["failed"]:
            print(f"  第 {row[0]} 行：{row[1]} / {row[2]} / {row[3]}")
        raise SystemExit(1)
    conn.close()


if __name__ == "__main__":
    main()
