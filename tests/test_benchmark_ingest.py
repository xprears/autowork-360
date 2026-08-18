# -*- coding: utf-8 -*-
"""基准汇总表导入与查询测试。"""

import tempfile
import unittest
from pathlib import Path

import openpyxl

from equipment.benchmark_ingest import ingest_benchmark_xlsx, rebuild_benchmark
from equipment.cli import query_benchmark_summary
from equipment.db import get_connection


class TestBenchmarkIngest(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.db = get_connection(Path(self.tmp.name) / "test.db")

    def tearDown(self):
        self.db.close()
        self.tmp.cleanup()

    def _make_benchmark_xlsx(self, path):
        """按基准表真实行号构造 3 个统计单位的迷你测试表。"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        districts = [
            ("天津市应急管理局", 72, 0, 1, 317, 0, 312, 10, 2, 2, 0, 5, 11),
            ("滨海新区", 39, 0, 1, 215, 420, 215, 315, 1, 1, 4, 9, 4),
            ("武清区", 32, 0, 9, 100, 0, 100, 213, 0, 2, 50, 3, 9),
        ]
        for offset, values in enumerate(districts):
            row_idx = 10 + offset
            ws.cell(row=row_idx, column=1, value=values[0])
            # 卫星电话手持（市局配发/区局自购）
            ws.cell(row=row_idx, column=2, value=values[1])
            ws.cell(row=row_idx, column=3, value=values[2])
            # 固移
            ws.cell(row=row_idx, column=4, value=values[3])
            # 370MHz（市局配发/区局自购/可用）
            ws.cell(row=row_idx, column=5, value=values[4])
            ws.cell(row=row_idx, column=6, value=values[5])
            ws.cell(row=row_idx, column=7, value=values[6])
            # 北斗
            ws.cell(row=row_idx, column=8, value=values[7])
            # 右侧面板：便携站/指挥车/叫应/布控球/无人机
            ws.cell(row=row_idx, column=10, value=values[8])
            ws.cell(row=row_idx, column=11, value=values[9])
            ws.cell(row=row_idx, column=12, value=values[10])
            ws.cell(row=row_idx, column=13, value=values[11])
            ws.cell(row=row_idx, column=14, value=values[12])
        wb.save(path)

    def test_ingest_benchmark(self):
        xlsx = Path(self.tmp.name) / "基准.xlsx"
        self._make_benchmark_xlsx(xlsx)
        stats = ingest_benchmark_xlsx(xlsx, self.db)

        self.assertEqual(stats["total"], 35)
        self.assertEqual(len(stats["districts"]), 3)

        drone_rows = self.db.execute(
            "SELECT district, count FROM equipment_benchmark"
            " WHERE device_type = '无人机' ORDER BY district"
        ).fetchall()
        self.assertEqual(len(drone_rows), 3)
        self.assertEqual(
            {r["district"]: r["count"] for r in drone_rows},
            {"天津市应急管理局": 11, "滨海新区": 4, "武清区": 9},
        )

        # 市局配发 / 区局自购、配置总数 / 可用数量均已保存
        m370 = self.db.execute(
            "SELECT metric, count FROM equipment_benchmark"
            " WHERE device_type = '370MHz手持终端' AND district = '滨海新区'"
            " ORDER BY metric"
        ).fetchall()
        self.assertEqual(
            {r["metric"]: r["count"] for r in m370},
            {"可用数量": 215, "市局配发": 215, "区局自购": 420, "配置总数": 635},
        )

    def test_reingest_no_duplicate_and_rebuild(self):
        xlsx = Path(self.tmp.name) / "基准.xlsx"
        self._make_benchmark_xlsx(xlsx)
        ingest_benchmark_xlsx(xlsx, self.db)
        ingest_benchmark_xlsx(xlsx, self.db)
        count = self.db.execute("SELECT COUNT(*) FROM equipment_benchmark").fetchone()[0]
        self.assertEqual(count, 35)

        rebuild_benchmark(self.db)
        count = self.db.execute("SELECT COUNT(*) FROM equipment_benchmark").fetchone()[0]
        self.assertEqual(count, 0)

    def test_query_benchmark_summary(self):
        xlsx = Path(self.tmp.name) / "基准.xlsx"
        self._make_benchmark_xlsx(xlsx)
        ingest_benchmark_xlsx(xlsx, self.db)

        rows = query_benchmark_summary(self.db, device_type="无人机")
        self.assertEqual(len(rows), 3)
        self.assertEqual(sum(r["cnt"] for r in rows), 24)

        wuqing = query_benchmark_summary(
            self.db, district="武清区", device_type="无人机", metric="总数"
        )
        self.assertEqual(len(wuqing), 1)
        self.assertEqual(wuqing[0]["cnt"], 9)


if __name__ == "__main__":
    unittest.main()
